import os
import logging
import time
import datetime
import tempfile
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters
from db import db
import stats

logger = logging.getLogger(__name__)

class BotInteractivo:
    def __init__(self):
        self.token = os.environ.get("TELEGRAM_BOT_TOKEN")
        self.owner_id = int(os.environ.get("AUTHORIZED_USER_ID", 0))
        self.last_command_ts = {}
        self.app = None
        self.add_process = {} # Diccionario para rastrear el estado del usuario en /agregar
        
        self.add_steps = [
            "MATRICULA", "NOMBRE", "ORG", "PROV", "MARCA", "MODELO",
            "PRECIO", "COSTO", "LITROS", "CO2", "ICAO24"
        ]

    def _is_admin(self, user_id):
        if user_id == self.owner_id: return True
        auth_users = [u['user_id'] for u in db.get_usuarios_autorizados()]
        return user_id in auth_users

    def _check_auth(self, update: Update):
        uid = update.effective_user.id
        if not self._is_admin(uid):
            logger.warning(f"[BOT] Intento de acceso no autorizado: {uid}")
            return False, True  # Not auth, IS unauthorized attempt
        
        now = time.time()
        last = self.last_command_ts.get(uid, 0)
        if now - last < 2:
            return False, False
        self.last_command_ts[uid] = now
        return True, False  # Auth OK, not unauthorized

    def _check_owner(self, update: Update):
        uid = update.effective_user.id
        return uid == self.owner_id

    # --- COMANDOS BÁSICOS ---

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        auth, unauth = self._check_auth(update)
        if unauth:
            await update.message.reply_text("⛔ No tienes acceso. Contactá al administrador via Twitter/X: @AarwenAirTracker")
            return
        if not auth: return
        msg = "👋 <b>Hola! Soy Aerobot.</b>\n\n"
        msg += "El sistema de monitoreo híbrido está activo.\n"
        msg += "Usa /comandos para ver todo lo que puedo hacer."
        await update.message.reply_html(msg)

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, skip_auth=False):
        if not skip_auth and not self._check_auth(update)[0]: return
        msg = "📊 <b>Comandos Disponibles:</b>\n\n"
        msg += "/aviones o /flota - Ver la flota monitoreada\n"
        msg += "/rastreando - Ver aviones en vuelo ahora\n"
        msg += "/ranking - Top 10 de gastos\n"
        msg += "/posiciones - Últimas posiciones detectadas\n"
        msg += "/stats &lt;matricula&gt; - Info detallada de un avión\n"
        msg += "/weekly - Reporte semanal\n"
        msg += "/informe - Excel de todos los vuelos\n"
        msg += "/informe &lt;mat&gt; - Excel de un avión\n"
        msg += "/estado - Verifica el estado del bot\n"
        msg += "/comandos - Muestra esta lista\n\n"
        msg += "👑 <b>Comandos de Administrador:</b>\n"
        msg += "/autorizar &lt;id&gt; &lt;nombre&gt; - Dar acceso\n"
        msg += "/desautorizar &lt;id&gt; - Quitar acceso\n"
        msg += "/usuarios - Ver autorizados\n"
        msg += "/deletestats &lt;matricula&gt; - Borrar historial\n"
        msg += "/agregar - Agregar aeronave manualmente\n"
        msg += "/quitar &lt;matricula&gt; - Dejar de seguir\n\n"
        msg += "🕵️ <b>Cazador Inteligente:</b>\n"
        msg += "/candidatas - Ver hallazgos del cazador\n"
        msg += "/reclutar &lt;callsign&gt; - Reclutar candidata\n"
        msg += "/reclutar &lt;existente&gt; + &lt;candidata&gt; - Merge hex\n"
        msg += "/cancelar - Cancelar proceso de carga\n"
        msg += "/ignorar &lt;callsign&gt; - Lista negra permanente\n"
        msg += "/designorar &lt;callsign&gt; - Sacar de lista negra\n"
        msg += "/ignorados - Ver lista negra\n\n"
        msg += "🐦 <b>Twitter:</b>\n"
        msg += "/tweetstatus - Estado de Twitter\n"
        msg += "/agregar_tweet &lt;mat&gt; - Autorizar Twitter\n"
        msg += "/quitar_tweet &lt;mat&gt; - Revocar Twitter\n"
        msg += "/listar_tweet - Matrículas autorizadas\n\n"
        msg += "🆔 /miid - Ver tu ID de Telegram"
        await update.message.reply_html(msg)

    async def get_flota(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        aeronaves = db.get_all_aeronaves(solo_activas=True)
        msg = "📋 <b>FLOTA MONITOREADA:</b>\n\n"
        for a in aeronaves:
            msg += f"• <code>{a['matricula']}</code> - {a['nombre']} ({a['organismo']})\n"
        await update.message.reply_html(msg)

    async def get_posiciones(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        posiciones = db.get_ultimas_posiciones(limite=20)
        if not posiciones:
            await update.message.reply_text("No hay posiciones recientes registradas.")
            return
        msg = "📍 <b>ÚLTIMAS POSICIONES:</b>\n\n"
        for p in posiciones:
            msg += f"• <b>{p['matricula']}</b>: {p['lat']:.3f}, {p['lon']:.3f} ({p['timestamp_utc'][-8:-3]} UTC)\n"
        await update.message.reply_html(msg)

    async def get_ranking(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        ranking = stats.get_ranking_gasto()
        if not ranking:
            await update.message.reply_text("No hay datos de gastos suficientes.")
            return
        msg = "💸 <b>TOP 10 GASTOS USD:</b>\n\n"
        for i, r in enumerate(ranking, 1):
            msg += f"{i}. <b>{r['matricula']}</b>: ${int(r['total_gastos_usd'] or 0):,} USD\n"
        await update.message.reply_html(msg)

    async def get_stats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args:
            await update.message.reply_text("Uso: /stats <matricula>")
            return
        mat = context.args[0].upper()
        s = stats.get_stats_aeronave(mat)
        if not s:
            await update.message.reply_text(f"No hay estadísticas para {mat}.")
            return
        msg = f"📊 <b>STATS: {mat}</b>\n\n"
        msg += f"Vuelos: {s['total_vuelos']}\n"
        msg += f"Gasto Est: ${int(s['total_costo_usd'] or 0):,} USD\n"
        msg += f"Distancia: {int(s['total_distancia_km'] or 0):,} km\n"
        msg += f"Combustible: {int(s['total_consumo_fuel_l'] or 0):,} L\n"
        msg += f"Emisiones: {int(s['total_emisiones_co2_kg'] or 0):,} kg CO2"
        await update.message.reply_html(msg)

    async def get_weekly(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        res = stats.get_resumen_semanal()
        if not res or not res['totales']:
            await update.message.reply_text("No hay datos de esta semana.")
            return
        t = res['totales']
        m = res['mas_gastador']
        msg = "📅 <b>REPORTE SEMANAL</b>\n\n"
        msg += f"Vuelos: {t.get('total_vuelos', 0)}\n"
        msg += f"Gasto Total: ${int(t.get('total_usd', 0) or 0):,} USD\n"
        msg += f"Distancia: {int(t.get('total_km', 0) or 0):,} km\n"
        if m:
            msg += f"\n🏆 <b>Más gastador:</b> {m['matricula']} (${int(m['usd']):,} USD)"
        await update.message.reply_html(msg)

    async def get_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        v_activos = db.get_conteo_detectados()
        msg = "🌐 <b>ESTADO DEL SISTEMA</b>\n\n"
        msg += f"Vuelos en radar vivo: <b>{v_activos}</b>\n"
        msg += f"Status: <b>ONLINE y MONITOREANDO</b>"
        await update.message.reply_html(msg)

    async def get_rastreando(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        vuelos = db.get_vuelos_activos()
        if not vuelos:
            await update.message.reply_text("🔇 No hay aviones rastreados en este momento.")
            return
        msg = "✈️ <b>AVIONES EN RADAR AHORA:</b>\n\n"
        now_utc = datetime.datetime.now(datetime.timezone.utc)
        for v in vuelos:
            dur = ""
            try:
                dt = datetime.datetime.fromisoformat(v['despegue_utc'])
                if dt.tzinfo is None: dt = dt.replace(tzinfo=datetime.timezone.utc)
                mins = int((now_utc - dt).total_seconds() / 60)
                dur = f" ({mins}min en vuelo)"
            except: pass
            
            # Check last update
            last_msg = ""
            try:
                last_upd_str = v.get('last_update')
                if last_upd_str:
                    last_upd = datetime.datetime.fromisoformat(last_upd_str)
                    if last_upd.tzinfo is None: last_upd = last_upd.replace(tzinfo=datetime.timezone.utc)
                    diff_mins = int((now_utc - last_upd).total_seconds() / 60)
                    last_arg = (last_upd - datetime.timedelta(hours=3)).strftime('%H:%M')
                    
                    if diff_mins > 15:
                        last_msg = f"\n    ⚠️ <i>Última detección: {last_arg} ARG ({diff_mins}min ago)</i>"
                    else:
                        last_msg = f"\n    📡 <i>Última detección: {last_arg} ARG ({diff_mins}min ago)</i>"
            except: pass

            msg += f"• <b>{v['matricula']}</b> desde {v.get('origen_nombre', '?')}{dur}{last_msg}\n"
        await update.message.reply_html(msg)

    async def mi_id(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(f"Tu ID es: {update.effective_user.id}")

    # --- COMANDOS ADMIN ---

    async def autorizar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_owner(update): 
            await update.message.reply_text("Solo el dueño puede hacer esto.")
            return
        if len(context.args) < 2:
            await update.message.reply_text("Uso: /autorizar <id> <nombre>")
            return
        try:
            uid = int(context.args[0])
            name = " ".join(context.args[1:])
            db.add_usuario_autorizado(uid, name)
            await update.message.reply_text(f"Usuario {name} ({uid}) autorizado.")
        except ValueError:
            await update.message.reply_text("El ID debe ser numérico.")

    async def desautorizar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_owner(update): return
        if not context.args:
            await update.message.reply_text("Uso: /desautorizar <id>")
            return
        try:
            uid = int(context.args[0])
            db.remove_usuario_autorizado(uid)
            await update.message.reply_text(f"Usuario {uid} desautorizado.")
        except ValueError:
            pass

    async def usuarios(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_owner(update): return
        users = db.get_usuarios_autorizados()
        if not users:
            await update.message.reply_text("No hay usuarios adicionales autorizados.")
            return
        msg = "👥 <b>USUARIOS AUTORIZADOS:</b>\n"
        for u in users:
            msg += f"• {u['username']} (<code>{u['user_id']}</code>)\n"
        await update.message.reply_html(msg)

    async def deletestats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args:
            await update.message.reply_text("Uso: /deletestats <matricula>")
            return
        mat = context.args[0].upper()
        db.delete_stats_aeronave(mat)
        await update.message.reply_text(f"Historial de {mat} borrado.")

    async def agregar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        uid = update.effective_user.id
        self.add_process[uid] = {"step": 0, "data": {}}
        await update.message.reply_text(f"📝 Nueva Aeronave\n1. {self.add_steps[0]}:")

    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        uid = update.effective_user.id
        if uid in self.add_process:
            del self.add_process[uid]
            await update.message.reply_text("❌ Proceso cancelado.")
        else:
            await update.message.reply_text("No hay ningún proceso activo para cancelar.")

    async def reclutar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        uid = update.effective_user.id
        if not context.args:
            await update.message.reply_text("Uso:\n/reclutar <callsign> — Reclutar candidata nueva\n/reclutar <existente> + <candidata> — Unificar hex de candidata en avión existente")
            return
        
        # Detectar modo MERGE: /reclutar TC-69 + HERCULES
        args_text = " ".join(context.args).upper()
        if "+" in args_text:
            parts = [p.strip() for p in args_text.split("+", 1)]
            if len(parts) != 2 or not parts[0] or not parts[1]:
                await update.message.reply_text("❌ Formato incorrecto. Uso: /reclutar <existente> + <candidata>")
                return
            
            mat_existente = parts[0]
            mat_candidata = parts[1]
            
            from db import db as db_instance
            with db_instance.connection_scope() as conn:
                cursor = db_instance.get_cursor(conn)
                ph = "%s" if db_instance.is_postgres else "?"
                
                # Verificar que el avión existente está en la base de datos
                avion = cursor.execute(f"SELECT matricula, nombre, icao24 FROM aeronaves WHERE matricula = {ph}", (mat_existente,)).fetchone()
                if not avion:
                    await update.message.reply_text(f"❌ No se encontró el avión '{mat_existente}' en la base de datos.")
                    return
                
                # Buscar la candidata y su hex
                candidata = cursor.execute(f"SELECT callsign, icao24 FROM aeronaves_candidatas WHERE callsign = {ph}", (mat_candidata,)).fetchone()
                if not candidata:
                    await update.message.reply_text(f"❌ No se encontró la candidata '{mat_candidata}'. Usá /candidatas para ver la lista.")
                    return
                
                nuevo_hex = candidata['icao24'] or ""
                if not nuevo_hex:
                    await update.message.reply_text(f"❌ La candidata '{mat_candidata}' no tiene código ICAO24 hex registrado.")
                    return
                
                viejo_hex = avion['icao24'] or "(vacío)"
                
                # Actualizar el hex del avión existente
                cursor.execute(f"UPDATE aeronaves SET icao24 = {ph} WHERE matricula = {ph}", (nuevo_hex, mat_existente))
                
                # Eliminar la candidata ya unificada
                cursor.execute(f"DELETE FROM aeronaves_candidatas WHERE callsign = {ph}", (mat_candidata,))
                
                conn.commit()
            
            await update.message.reply_html(
                f"✅ <b>MERGE COMPLETADO</b>\n\n"
                f"✈️ Avión: <b>{avion['nombre']}</b> ({mat_existente})\n"
                f"🔄 Hex anterior: <code>{viejo_hex}</code>\n"
                f"🆕 Hex nuevo: <code>{nuevo_hex}</code> (de candidata '{mat_candidata}')\n"
                f"🗑️ Candidata '{mat_candidata}' eliminada de la lista."
            )
            logger.info(f"[BOT] MERGE: {mat_existente} actualizado con hex {nuevo_hex} desde candidata {mat_candidata}")
            return
        
        # --- Modo normal: reclutar candidata nueva ---
        callsign = context.args[0].upper()
        # Buscar en candidatas
        from db import db as db_instance
        with db_instance.connection_scope() as conn:
            cursor = db_instance.get_cursor(conn)
            row = cursor.execute("SELECT callsign, icao24 FROM aeronaves_candidatas WHERE callsign = %s" if db_instance.is_postgres else "SELECT callsign, icao24 FROM aeronaves_candidatas WHERE callsign = ?", (callsign,)).fetchone()
            
            if not row:
                await update.message.reply_text(f"❌ No se encontró la candidata '{callsign}'. Usá /candidatas para ver la lista.")
                return
                
            icao = row['icao24'] or ""
            
        self.add_process[uid] = {
            "step": 0, 
            "data": {"MATRICULA": callsign, "ICAO24": icao},
            "prefilled": ["MATRICULA", "ICAO24"],
            "is_reclutar": True
        }
        
        # Avanzar al primer paso no prellenado
        await self._next_add_step(update, uid)

    async def _next_add_step(self, update: Update, uid: int):
        state = self.add_process[uid]
        while state["step"] < len(self.add_steps) and self.add_steps[state["step"]] in state.get("prefilled", []):
            state["step"] += 1
            
        step = state["step"]
        if step < len(self.add_steps):
            await update.message.reply_text(f"{step + 1}. {self.add_steps[step]}:")
        else:
            # Terminamos, procesar y guardar
            d = state["data"]
            try:
                def parse_float(val):
                    try:
                        v_str = str(val).strip().lower()
                        if v_str in ["none", "noe", "", "na", "n/a", "no", "desconocido"]:
                            return 0.0
                        return float(v_str.replace(",", "."))
                    except ValueError:
                        return 0.0

                prov = d["PROV"]
                org = d["ORG"].lower()
                
                # Inteligencia de categoría auto-detectada únicamente si se inició el proceso via /reclutar
                if state.get("is_reclutar"):
                    # 1. Auto-detección Militar / Fuerzas de seguridad
                    palabras_militares = ["fuerza aerea", "policia", "prefectura", "gendarmeria", "ejercito", "armada", "militar", "seguridad"]
                    # 2. Auto-detección Privados (si contiene privado, empresa, persona)
                    palabras_privadas = ["privado", "empresa", "persona"]
                    
                    if any(p in org for p in palabras_militares):
                        prov = "Militar"
                    elif any(p in org for p in palabras_privadas):
                        prov = "Privado"
                    # 3. Si no pone 'nacion' o 'presidencia' en Org y no es militar ni privado, entonces es provincial
                    elif not any(p in org for p in ["nacion", "presidencia", "nacional"]):
                        if prov.lower() in ["nacional", "nacion", "", "none"]:
                            prov = "Provincial"

                data = {
                    "matricula": d["MATRICULA"].upper(),
                    "nombre": d["NOMBRE"],
                    "organismo": d["ORG"],
                    "provincia": prov,
                    "marca": d["MARCA"],
                    "modelo": d["MODELO"],
                    "precio_usd": parse_float(d["PRECIO"]),
                    "costo_hora_usd": parse_float(d["COSTO"]),
                    "litros_hora_estimado": parse_float(d["LITROS"]),
                    "co2_kg_hora_estimado": parse_float(d["CO2"]),
                    "icao24": d["ICAO24"].lower() if d["ICAO24"].lower() not in ["none", "noe", ""] else "",
                    "activa": 1
                }
                from db import db as db_instance
                db_instance.upsert_aeronave(data)
                
                # Borrar de candidatas si era un reclutamiento
                if "prefilled" in state:
                    with db_instance.connection_scope() as conn:
                        cursor = db_instance.get_cursor(conn)
                        cursor.execute("DELETE FROM aeronaves_candidatas WHERE callsign = %s" if db_instance.is_postgres else "DELETE FROM aeronaves_candidatas WHERE callsign = ?", (data["matricula"],))
                        
                await update.message.reply_text(f"✅ Aeronave {data['matricula']} guardada y activa en el radar.")
            except Exception as e:
                logger.error(f"[BOT] Error guardando aeronave: {e}")
                await update.message.reply_text("❌ Error interno al guardar la aeronave.")
            finally:
                del self.add_process[uid]

    async def handle_text(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        uid = update.effective_user.id
        
        if not update.message.text: return
        text = update.message.text.strip()
        
        if uid in self.add_process:
            if text.lower() == "/cancelar":
                del self.add_process[uid]
                await update.message.reply_text("❌ Proceso cancelado.")
                return
                
            state = self.add_process[uid]
            step = state["step"]
            field_name = self.add_steps[step]
            
            # Guardar el dato
            state["data"][field_name] = text
            
            # Siguiente paso
            state["step"] += 1
            await self._next_add_step(update, uid)
            
        else:
            await self.handle_unknown_command(update, context, skip_auth=True)


    async def quitar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args:
            await update.message.reply_text("Uso: /quitar <matricula>")
            return
        mat = context.args[0].upper()
        db.update_aeronave_status(mat, 0)
        await update.message.reply_text(f"Aeronave {mat} desactivada del rastreo.")

    async def candidatas(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        cand = db.get_aeronaves_candidatas()
        if not cand:
            await update.message.reply_text("No hay candidatas detectadas por el cazador.")
            return
        msg = "🕵️‍♂️ <b>CANDIDATAS DEL CAZADOR:</b>\n\n"
        for c in cand[:20]:
            msg += f"• <code>{c['callsign']}</code> ({c['icao24']}) - Visto {c['veces_visto']} veces\n"
        await update.message.reply_html(msg)

    async def ignorar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args:
            await update.message.reply_text("Uso: /ignorar <callsign>\nEjemplo: /ignorar FAU595")
            return
        callsign = context.args[0].upper()
        db.add_callsign_ignorado(callsign)
        await update.message.reply_html(f"🚫 <code>{callsign}</code> agregado a la lista negra.\nYa no aparecerá en /candidatas nunca más.")

    async def designorar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args:
            await update.message.reply_text("Uso: /designorar <callsign>")
            return
        callsign = context.args[0].upper()
        removed = db.remove_callsign_ignorado(callsign)
        if removed:
            await update.message.reply_html(f"✅ <code>{callsign}</code> removido de la lista negra.\nVolverá a aparecer en /candidatas si el cazador lo detecta.")
        else:
            await update.message.reply_text(f"'{callsign}' no estaba en la lista negra.")

    async def ignorados(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        lista = db.get_callsigns_ignorados()
        if not lista:
            await update.message.reply_text("No hay callsigns ignorados.")
            return
        msg = "🚫 <b>CALLSIGNS IGNORADOS:</b>\n\n"
        for item in lista:
            msg += f"• <code>{item['callsign']}</code> (desde {item['fecha'][:10]})\n"
        await update.message.reply_html(msg)

    async def tweetstatus(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        enabled = os.environ.get("TWITTER_ENABLED") == "1"
        wl = db.get_setting("twitter_whitelist", "")
        msg = f"🐦 Twitter Enabled: <b>{'SI' if enabled else 'NO'}</b>\n"
        msg += f"Filtro activo para: {wl if wl else 'TODAS'}"
        await update.message.reply_html(msg)

    async def agregar_tweet(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args: return
        mat = context.args[0].upper()
        wl = db.get_setting("twitter_whitelist", "")
        mats = [m.strip() for m in wl.split(",")] if wl else []
        if mat not in mats:
            mats.append(mat)
            db.set_setting("twitter_whitelist", ",".join(mats))
        await update.message.reply_text(f"{mat} agregada a la whitelist de Twitter.")

    async def quitar_tweet(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        if not context.args: return
        mat = context.args[0].upper()
        wl = db.get_setting("twitter_whitelist", "")
        mats = [m.strip() for m in wl.split(",")] if wl else []
        if mat in mats:
            mats.remove(mat)
            db.set_setting("twitter_whitelist", ",".join(mats))
        await update.message.reply_text(f"{mat} removida de la whitelist de Twitter.")

    async def listar_tweet(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        wl = db.get_setting("twitter_whitelist", "")
        await update.message.reply_text(f"Whitelist Twitter: {wl if wl else 'Vacia (Aplica a todas)'}")

    async def generar_informe(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self._check_auth(update)[0]: return
        
        await update.message.reply_text("⏳ Generando informe Excel, espere...")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Determinar si es informe general o por matrícula
            mat_filter = context.args[0].upper() if context.args else None
            
            if mat_filter:
                vuelos = db.get_vuelos_por_matricula(mat_filter)
                filename = f"informe_{mat_filter}.xlsx"
                titulo = f"Informe de Vuelos - {mat_filter}"
            else:
                vuelos = db.get_vuelos_historial()
                filename = "informe_flota_completa.xlsx"
                titulo = "Informe de Vuelos - Flota Completa"
            
            if not vuelos:
                await update.message.reply_text(f"❌ No hay vuelos registrados{' para ' + mat_filter if mat_filter else ''}.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Vuelos"
            
            # Estilos
            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
            
            # Título
            ws.merge_cells('A1:L1')
            ws['A1'] = titulo
            ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:L2')
            ws['A2'] = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} - Total: {len(vuelos)} vuelos"
            ws['A2'].font = Font(name='Calibri', size=10, italic=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Encabezados
            headers = ['Fecha', 'Matrícula', 'Nombre', 'Organismo', 'Modelo',
                       'Origen', 'Destino', 'Despegue (ARG)', 'Aterrizaje (ARG)',
                       'Duración', 'Costo USD', 'Internacional']
            
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Datos
            for idx, v in enumerate(vuelos):
                row = idx + 5
                
                # Parsear fechas UTC a ARG (-3)
                try:
                    dt_desp = datetime.datetime.fromisoformat(v.get('despegue_utc', ''))
                    desp_arg = (dt_desp - datetime.timedelta(hours=3)).strftime('%d/%m/%Y')
                    hora_desp = (dt_desp - datetime.timedelta(hours=3)).strftime('%H:%M')
                except:
                    desp_arg = v.get('despegue_utc', '?')
                    hora_desp = '?'
                
                try:
                    dt_aterr = datetime.datetime.fromisoformat(v.get('aterrizaje_utc', ''))
                    hora_aterr = (dt_aterr - datetime.timedelta(hours=3)).strftime('%H:%M')
                except:
                    hora_aterr = '?'
                
                dur = v.get('duracion_min', 0) or 0
                horas = dur // 60
                minutos = dur % 60
                dur_str = f"{horas}h {minutos}min" if horas > 0 else f"{minutos}min"
                
                origen = f"{v.get('origen_nombre', '?')} ({v.get('origen_ciudad', '?')})"
                destino = f"{v.get('destino_nombre', '?')} ({v.get('destino_ciudad', '?')})"
                
                row_data = [
                    desp_arg,
                    v.get('matricula', '?'),
                    v.get('nombre', '?'),
                    v.get('organismo', '?'),
                    v.get('modelo', '?'),
                    origen,
                    destino,
                    hora_desp,
                    hora_aterr,
                    dur_str,
                    v.get('costo_usd', 0),
                    'Sí' if v.get('es_internacional') else 'No'
                ]
                
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.font = Font(name='Calibri', size=10)
                    if idx % 2 == 1:
                        cell.fill = alt_fill
                    if col == 11 and isinstance(val, (int, float)):
                        cell.number_format = '#,##0'
            
            # Ajustar anchos
            col_widths = [12, 12, 20, 18, 14, 28, 28, 14, 14, 12, 12, 14]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64+i)].width = w
            
            # Guardar y enviar
            filepath = os.path.join(tempfile.gettempdir(), filename)
            wb.save(filepath)
            
            with open(filepath, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=filename,
                    caption=f"📊 {titulo}\n✈️ {len(vuelos)} vuelos registrados"
                )
            
            os.remove(filepath)
            logger.info(f"[BOT] Informe Excel generado: {filename} ({len(vuelos)} vuelos)")
            
        except ImportError:
            await update.message.reply_text("❌ Error: openpyxl no está instalado. Ejecutar: pip install openpyxl")
        except Exception as e:
            logger.error(f"[BOT] Error generando informe: {e}")
            await update.message.reply_text(f"❌ Error generando informe: {str(e)}")

    async def run_async(self):
        if not self.token:
            logger.error("[BOT] No hay TOKEN de Telegram configurado.")
            return
        
        self.app = ApplicationBuilder().token(self.token).build()
        
        # Básicos
        self.app.add_handler(CommandHandler("start", self.start))
        self.app.add_handler(CommandHandler("help", self.help_command))
        self.app.add_handler(CommandHandler("comandos", self.help_command))
        self.app.add_handler(CommandHandler("aviones", self.get_flota))
        self.app.add_handler(CommandHandler("flota", self.get_flota))
        self.app.add_handler(CommandHandler("ranking", self.get_ranking))
        self.app.add_handler(CommandHandler("posiciones", self.get_posiciones))
        self.app.add_handler(CommandHandler("pos", self.get_posiciones))
        self.app.add_handler(CommandHandler("stats", self.get_stats))
        self.app.add_handler(CommandHandler("weekly", self.get_weekly))
        self.app.add_handler(CommandHandler("estado", self.get_status))
        self.app.add_handler(CommandHandler("status", self.get_status))
        self.app.add_handler(CommandHandler("miid", self.mi_id))

        self.app.add_handler(CommandHandler("rastreando", self.get_rastreando))

        # Admin
        self.app.add_handler(CommandHandler("autorizar", self.autorizar))
        self.app.add_handler(CommandHandler("desautorizar", self.desautorizar))
        self.app.add_handler(CommandHandler("usuarios", self.usuarios))
        self.app.add_handler(CommandHandler("deletestats", self.deletestats))
        self.app.add_handler(CommandHandler("agregar", self.agregar))
        self.app.add_handler(CommandHandler("reclutar", self.reclutar))
        self.app.add_handler(CommandHandler("cancelar", self.cancelar))
        self.app.add_handler(CommandHandler("quitar", self.quitar))
        self.app.add_handler(CommandHandler("candidatas", self.candidatas))
        self.app.add_handler(CommandHandler("ignorar", self.ignorar))
        self.app.add_handler(CommandHandler("designorar", self.designorar))
        self.app.add_handler(CommandHandler("ignorados", self.ignorados))
        self.app.add_handler(CommandHandler("tweetstatus", self.tweetstatus))
        self.app.add_handler(CommandHandler("agregar_tweet", self.agregar_tweet))
        self.app.add_handler(CommandHandler("quitar_tweet", self.quitar_tweet))
        self.app.add_handler(CommandHandler("listar_tweet", self.listar_tweet))
        self.app.add_handler(CommandHandler("informe", self.generar_informe))

        # Message handler para el flujo conversacional de agregar
        self.app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), self.handle_text))
        
        # Handler de comandos no reconocidos (DEBE IR ÚLTIMO)
        self.app.add_handler(MessageHandler(filters.COMMAND, self.handle_unknown_command))

        logger.info("[BOT] Iniciando Polling...")
        await self.app.initialize()
        await self.app.start()
        await self.app.updater.start_polling()

    async def handle_unknown_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, skip_auth=False):
        if not skip_auth:
            auth, unauth = self._check_auth(update)
            if unauth:
                await update.message.reply_text("⛔ No tienes acceso. Contactá al administrador via Twitter/X: @AarwenAirTracker")
                return
            if not auth: return
            
        await update.message.reply_text("❓ Comando inválido o texto no reconocido. Los comandos válidos son:")
        await self.help_command(update, context, skip_auth=True)
